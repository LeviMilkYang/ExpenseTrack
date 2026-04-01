from __future__ import annotations

import argparse
import ast
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DEFAULT_SOURCE_PATH = PROJECT_DIR / "expense.xlsx"
DEFAULT_REPORT_PATH = PROJECT_DIR / "expense_report.xlsx"

DATA_HEADERS = ["Date", "Time", "Timezone", "Amount", "Currency", "Type", "Category", "Note", "Status"]
DATA_HEADERS_WITH_ID = ["ID", "Date", "Time", "Timezone", "Amount", "Currency", "Type", "Category", "Note", "Status"]
LEGACY_DATA_HEADERS = ["Date", "Time", "Amount", "Currency", "Type", "Category", "Note", "Status"]
BAR_HEADERS = ["期间", "收入合计", "支出合计", "结余合计"]
PIE_HEADERS = ["分类", "支出合计"]
LOAN_TYPES = {"借入", "贷出", "收回", "偿还"}
VOID_MARK = "作废"


@dataclass(frozen=True)
class LedgerRecord:
    sheet_name: str
    record_date: date
    amount: Decimal
    currency: str
    record_type: str
    category: str
    need_confirm: str


@dataclass(frozen=True)
class HeaderLayout:
    date_idx: int
    amount_idx: int
    currency_idx: int
    type_idx: int
    category_idx: int
    status_idx: int


class _FormulaEvaluator(ast.NodeVisitor):
    allowed_binops = {
        ast.Add: lambda a, b: a + b,
        ast.Sub: lambda a, b: a - b,
        ast.Mult: lambda a, b: a * b,
        ast.Div: lambda a, b: a / b,
    }

    allowed_unaryops = {
        ast.UAdd: lambda a: a,
        ast.USub: lambda a: -a,
    }

    def visit_Expression(self, node: ast.Expression) -> Decimal:
        return self.visit(node.body)

    def visit_BinOp(self, node: ast.BinOp) -> Decimal:
        op_type = type(node.op)
        if op_type not in self.allowed_binops:
            raise ValueError(f"unsupported operator: {op_type.__name__}")
        return self.allowed_binops[op_type](self.visit(node.left), self.visit(node.right))

    def visit_UnaryOp(self, node: ast.UnaryOp) -> Decimal:
        op_type = type(node.op)
        if op_type not in self.allowed_unaryops:
            raise ValueError(f"unsupported operator: {op_type.__name__}")
        return self.allowed_unaryops[op_type](self.visit(node.operand))

    def visit_Constant(self, node: ast.Constant) -> Decimal:
        if not isinstance(node.value, (int, float)):
            raise ValueError(f"unsupported literal: {node.value!r}")
        return Decimal(str(node.value))

    def generic_visit(self, node: ast.AST) -> Decimal:
        raise ValueError(f"unsupported expression: {type(node).__name__}")


def _to_decimal(value) -> Decimal | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("="):
        parsed = ast.parse(text[1:], mode="eval")
        return _FormulaEvaluator().visit(parsed)
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"金额无法转换为数字: {value!r}") from exc


def _normalize_amount(value: Decimal) -> int | float:
    quantized = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if quantized == quantized.to_integral():
        return int(quantized)
    return float(quantized)


def _parse_excel_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    return datetime.strptime(text, "%Y-%m-%d").date()


def _detect_layout(sheet) -> HeaderLayout | None:
    headers_with_id = [sheet.cell(row=1, column=idx).value for idx in range(1, len(DATA_HEADERS_WITH_ID) + 1)]
    if headers_with_id == DATA_HEADERS_WITH_ID:
        return HeaderLayout(date_idx=1, amount_idx=4, currency_idx=5, type_idx=6, category_idx=7, status_idx=9)

    headers = [sheet.cell(row=1, column=idx).value for idx in range(1, len(DATA_HEADERS) + 1)]
    if headers == DATA_HEADERS:
        return HeaderLayout(date_idx=0, amount_idx=3, currency_idx=4, type_idx=5, category_idx=6, status_idx=8)

    legacy_headers = [sheet.cell(row=1, column=idx).value for idx in range(1, len(LEGACY_DATA_HEADERS) + 1)]
    if legacy_headers == LEGACY_DATA_HEADERS:
        return HeaderLayout(date_idx=0, amount_idx=2, currency_idx=3, type_idx=4, category_idx=5, status_idx=7)

    return None


def _load_records(source_path: Path) -> list[LedgerRecord]:
    workbook = load_workbook(source_path, data_only=False)
    records: list[LedgerRecord] = []

    for sheet_name in workbook.sheetnames:
        if not sheet_name.isdigit():
            continue

        sheet = workbook[sheet_name]
        layout = _detect_layout(sheet)
        if layout is None:
            continue

        max_col = max(
            layout.date_idx,
            layout.amount_idx,
            layout.currency_idx,
            layout.type_idx,
            layout.category_idx,
            layout.status_idx,
        ) + 1
        for row in sheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
            if not any(value not in (None, "") for value in row):
                continue

            record_type = str(row[layout.type_idx]).strip()
            if record_type not in {"收入", "支出"}:
                continue

            status = "" if row[layout.status_idx] is None else str(row[layout.status_idx]).strip()
            if status == VOID_MARK:
                continue

            amount = _to_decimal(row[layout.amount_idx])
            if amount is None:
                continue

            records.append(
                LedgerRecord(
                    sheet_name=sheet_name,
                    record_date=_parse_excel_date(row[layout.date_idx]),
                    amount=amount,
                    currency=str(row[layout.currency_idx]).strip() or "CNY",
                    record_type=record_type,
                    category=str(row[layout.category_idx]).strip(),
                    need_confirm=status,
                )
            )
    return records


def _filter_currency(records: Iterable[LedgerRecord], currency: str) -> list[LedgerRecord]:
    return [record for record in records if record.currency == currency and record.record_type not in LOAN_TYPES]


def _apply_cutoff(records: Iterable[LedgerRecord], cutoff_date: date | None) -> list[LedgerRecord]:
    if cutoff_date is None:
        return list(records)
    return [record for record in records if record.record_date <= cutoff_date]


def _build_monthly_rows(records: Iterable[LedgerRecord]) -> list[list[object]]:
    summary: dict[str, dict[str, Decimal]] = defaultdict(lambda: {"收入": Decimal("0"), "支出": Decimal("0")})
    for record in records:
        month_key = record.record_date.strftime("%Y-%m")
        summary[month_key][record.record_type] += record.amount

    rows = []
    for month_key in sorted(summary):
        income = summary[month_key]["收入"]
        expense = summary[month_key]["支出"]
        rows.append([month_key, _normalize_amount(income), _normalize_amount(expense), _normalize_amount(income - expense)])
    return rows


def _build_yearly_rows(records: Iterable[LedgerRecord]) -> list[list[object]]:
    summary: dict[str, dict[str, Decimal]] = defaultdict(lambda: {"收入": Decimal("0"), "支出": Decimal("0")})
    for record in records:
        year_key = str(record.record_date.year)
        summary[year_key][record.record_type] += record.amount

    rows = []
    for year_key in sorted(summary):
        income = summary[year_key]["收入"]
        expense = summary[year_key]["支出"]
        rows.append([year_key, _normalize_amount(income), _normalize_amount(expense), _normalize_amount(income - expense)])
    return rows


def _build_category_rows(records: Iterable[LedgerRecord]) -> tuple[list[list[object]], dict[str, list[list[object]]]]:
    total_summary: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    yearly_summary: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal("0")))

    for record in records:
        if record.record_type != "支出":
            continue
        total_summary[record.category] += record.amount
        yearly_summary[str(record.record_date.year)][record.category] += record.amount

    total_rows = [
        [category, _normalize_amount(amount)]
        for category, amount in sorted(total_summary.items(), key=lambda item: (-item[1], item[0]))
    ]
    yearly_rows = {
        year: [[category, _normalize_amount(amount)] for category, amount in sorted(values.items(), key=lambda item: (-item[1], item[0]))]
        for year, values in sorted(yearly_summary.items())
    }
    return total_rows, yearly_rows


def _autosize(sheet) -> None:
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for column_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(column_idx)].width = min(max(width + 2, 12), 24)


def _style_table(sheet, headers: list[str], start_row: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=index, value=header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=start_row + 1):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00;-#,##0.00'
    _autosize(sheet)


def _summary_totals(rows: list[list[object]]) -> tuple[float, float, float]:
    income = sum(float(row[1]) for row in rows)
    expense = sum(float(row[2]) for row in rows)
    return income, expense, income - expense


def _write_kpi_card(sheet, label_cell: str, value_cell: str, label: str, value: float | str) -> None:
    sheet[label_cell] = label
    sheet[label_cell].font = Font(bold=True, color="FFFFFF")
    sheet[label_cell].fill = PatternFill("solid", fgColor="1F4E78")
    sheet[label_cell].alignment = Alignment(horizontal="center")

    sheet[value_cell] = value
    sheet[value_cell].font = Font(size=13, bold=True)
    sheet[value_cell].fill = PatternFill("solid", fgColor="DCE6F1")
    sheet[value_cell].alignment = Alignment(horizontal="center")
    if isinstance(value, (int, float)):
        sheet[value_cell].number_format = '#,##0.00;-#,##0.00'


def _build_summary_sheet(
    workbook: Workbook,
    currency: str,
    monthly_rows: list[list[object]],
    yearly_rows: list[list[object]],
    as_of_date: date | None,
) -> None:
    sheet = workbook.create_sheet(title=f"总览_{currency}")
    sheet.freeze_panes = None
    sheet["A1"] = f"{currency} 收支统计"
    sheet["A1"].font = Font(size=16, bold=True)
    cutoff_month = as_of_date.strftime("%Y-%m") if as_of_date else "无数据"
    cutoff_date_text = as_of_date.strftime("%Y-%m-%d") if as_of_date else "无数据"
    sheet["A2"] = f"统计截止月份：{cutoff_month}（截至 {cutoff_date_text}）"
    sheet["A4"] = "统计口径：仅统计收入/支出，排除借贷操作与已作废记录。"

    total_income, total_expense, total_balance = _summary_totals(monthly_rows)
    _write_kpi_card(sheet, "A6", "B6", "累计收入", total_income)
    _write_kpi_card(sheet, "C6", "D6", "累计支出", total_expense)
    _write_kpi_card(sheet, "E6", "F6", "累计结余", total_balance)

    start_month_row = 9
    for offset, header in enumerate(BAR_HEADERS, start=1):
        sheet.cell(row=start_month_row, column=offset, value=header)
    for row_offset, values in enumerate(monthly_rows, start=1):
        for column_idx, value in enumerate(values, start=1):
            sheet.cell(row=start_month_row + row_offset, column=column_idx, value=value)
    _style_table(sheet, BAR_HEADERS, start_month_row)

    start_year_row = start_month_row + len(monthly_rows) + 4
    for offset, header in enumerate(BAR_HEADERS, start=1):
        sheet.cell(row=start_year_row, column=offset, value=header)
    for row_offset, values in enumerate(yearly_rows, start=1):
        for column_idx, value in enumerate(values, start=1):
            sheet.cell(row=start_year_row + row_offset, column=column_idx, value=value)

    monthly_chart = BarChart()
    monthly_chart.type = "col"
    monthly_chart.style = 10
    monthly_chart.title = "逐月收入 / 支出 / 结余"
    monthly_chart.y_axis.title = f"金额 ({currency})"
    monthly_chart.x_axis.title = "月份"
    monthly_chart.height = 9
    monthly_chart.width = 16
    if monthly_rows:
        data = Reference(sheet, min_col=2, max_col=4, min_row=start_month_row, max_row=start_month_row + len(monthly_rows))
        categories = Reference(sheet, min_col=1, min_row=start_month_row + 1, max_row=start_month_row + len(monthly_rows))
        monthly_chart.add_data(data, titles_from_data=True)
        monthly_chart.set_categories(categories)
        monthly_chart.legend.position = "r"
    sheet.add_chart(monthly_chart, "H3")

    yearly_chart = BarChart()
    yearly_chart.type = "col"
    yearly_chart.style = 11
    yearly_chart.title = "逐年收入 / 支出 / 结余"
    yearly_chart.y_axis.title = f"金额 ({currency})"
    yearly_chart.x_axis.title = "年份"
    yearly_chart.height = 9
    yearly_chart.width = 16
    if yearly_rows:
        data = Reference(sheet, min_col=2, max_col=4, min_row=start_year_row, max_row=start_year_row + len(yearly_rows))
        categories = Reference(sheet, min_col=1, min_row=start_year_row + 1, max_row=start_year_row + len(yearly_rows))
        yearly_chart.add_data(data, titles_from_data=True)
        yearly_chart.set_categories(categories)
        yearly_chart.legend.position = "r"
    chart_row = max(start_year_row + len(yearly_rows) + 3, 22)
    sheet.add_chart(yearly_chart, f"H{chart_row}")

    _style_table(sheet, BAR_HEADERS, start_year_row)
    _autosize(sheet)


def _add_pie_chart(sheet, title: str, start_row: int, chart_cell: str, rows: list[list[object]]) -> None:
    sheet.cell(row=start_row, column=1, value=title).font = Font(size=13, bold=True)
    for offset, header in enumerate(PIE_HEADERS, start=1):
        cell = sheet.cell(row=start_row + 1, column=offset, value=header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row_offset, values in enumerate(rows, start=1):
        for column_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=start_row + 1 + row_offset, column=column_idx, value=value)
            if column_idx == 2:
                cell.number_format = '#,##0.00;-#,##0.00'

    if not rows:
        sheet.cell(row=start_row + 2, column=1, value="暂无支出数据")
        return

    chart = PieChart()
    chart.style = 10
    chart.title = title
    chart.height = 8
    chart.width = 11
    labels = Reference(sheet, min_col=1, min_row=start_row + 2, max_row=start_row + 1 + len(rows))
    data = Reference(sheet, min_col=2, min_row=start_row + 1, max_row=start_row + 1 + len(rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showLeaderLines = True
    sheet.add_chart(chart, chart_cell)


def _build_category_sheet(workbook: Workbook, currency: str, total_rows: list[list[object]], yearly_rows: dict[str, list[list[object]]]) -> None:
    sheet = workbook.create_sheet(title=f"分类占比_{currency}")
    sheet.freeze_panes = None
    sheet["A1"] = f"{currency} 各分类支出占比"
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A3"] = "统计口径：仅统计支出，排除借贷操作与已作废记录。"

    _add_pie_chart(sheet, "全部年份支出占比", 5, "D5", total_rows)
    current_row = max(5 + len(total_rows) + 4, 22)
    chart_slots = ["D", "L"]
    slot_index = 0
    for year, rows in yearly_rows.items():
        chart_column = f"{chart_slots[slot_index % len(chart_slots)]}{current_row}"
        _add_pie_chart(sheet, f"{year} 年支出占比", current_row, chart_column, rows)
        if slot_index % 2 == 1:
            current_row += max(len(rows) + 16, 18)
        slot_index += 1

    _autosize(sheet)


def refresh_report_workbook(source_path: str | Path, report_path: str | Path | None = None, cutoff_date: date | None = None) -> Path:
    source = Path(source_path).resolve()
    target = Path(report_path).resolve() if report_path else source.with_name("expense_report.xlsx")
    records = _apply_cutoff(_load_records(source), cutoff_date)
    currencies = sorted({record.currency for record in records}) or ["CNY"]
    as_of_date = max((record.record_date for record in records), default=None)

    workbook = Workbook()
    workbook.remove(workbook.active)

    index_sheet = workbook.create_sheet(title="说明")
    index_sheet["A1"] = "Expense Report"
    index_sheet["A1"].font = Font(size=18, bold=True)
    index_sheet["A3"] = f"源文件：{source.name}"
    index_sheet["A4"] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    index_sheet["A6"] = "规则："
    index_sheet["A7"] = "1. 仅统计收入/支出。"
    index_sheet["A8"] = "2. 借入、贷出、收回、偿还不计入统计。"
    index_sheet["A9"] = "3. Status=作废 的记录不计入统计。"
    index_sheet["A10"] = "4. 报表是独立工作簿，不会修改原始记账数据。"
    _autosize(index_sheet)

    for currency in currencies:
        currency_records = _filter_currency(records, currency)
        monthly_rows = _build_monthly_rows(currency_records)
        yearly_rows = _build_yearly_rows(currency_records)
        total_rows, yearly_category_rows = _build_category_rows(currency_records)
        _build_summary_sheet(workbook, currency, monthly_rows, yearly_rows, as_of_date)
        _build_category_sheet(workbook, currency, total_rows, yearly_category_rows)

    workbook.save(target)
    return target


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate chart-based report workbook from expense.xlsx.")
    parser.add_argument("--source", default=str(DEFAULT_SOURCE_PATH))
    parser.add_argument("--report", default=str(DEFAULT_REPORT_PATH))
    parser.add_argument("--cutoff-date", default="", help="Only include records on or before YYYY-MM-DD.")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    cutoff_date = datetime.strptime(args.cutoff_date, "%Y-%m-%d").date() if str(args.cutoff_date).strip() else None
    report_path = refresh_report_workbook(args.source, args.report, cutoff_date=cutoff_date)
    print(report_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
