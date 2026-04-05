from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DEFAULT_EXCEL_PATH = PROJECT_DIR / "expense.xlsx"
HELPER_TEMP_DIR = BASE_DIR / "tmp_excel_helper"
DEFAULT_CONFIG_PATH = BASE_DIR / "telegram_bot_config.json"
FALLBACK_CATEGORY = "未分类"

STATUS_NORMAL = ""
STATUS_PENDING = "待确认"
STATUS_VOID = "作废"
ALLOWED_STATUS = {STATUS_NORMAL, STATUS_PENDING, STATUS_VOID}
DEFAULT_TIMEZONE = "UTC+08:00"
TIMEZONE_PATTERN = re.compile(r"^(?:UTC|GMT)\s*([+-])\s*(\d{1,2})(?::?(\d{2}))?$", re.IGNORECASE)

# 新增 ID 列在首位
EXPECTED_HEADERS = [
    "ID",
    "Date",
    "Time",
    "Timezone",
    "Amount",
    "Currency",
    "Type",
    "Category",
    "Note",
    "PaymentChannel",
    "Status",
]

FIELD_ALIASES = {
    "ID": ("ID", "id"),
    "Date": ("Date", "date", "日期"),
    "Time": ("Time", "time", "时间"),
    "Timezone": ("Timezone", "timezone", "tz", "时区"),
    "Amount": ("Amount", "amount", "金额"),
    "Currency": ("Currency", "currency", "币种"),
    "Type": ("Type", "type", "收支"),
    "Category": ("Category", "category", "分类"),
    "Note": ("Note", "note", "备注"),
    "PaymentChannel": ("PaymentChannel", "payment_channel", "支付渠道"),
    "Status": ("Status", "status", "状态", "NeedConfirm", "need_confirm", "待确认"),
}

TYPE_ALIASES = {
    "收入": "收入",
    "收": "收入",
    "income": "收入",
    "支出": "支出",
    "支": "支出",
    "expense": "支出",
    "借入": "借入",
    "借到": "借入",
    "借款": "借入",
    "borrow": "借入",
    "loanin": "借入",
    "贷出": "贷出",
    "借出": "贷出",
    "出借": "贷出",
    "lend": "贷出",
    "loanout": "贷出",
    "还入": "收回",
    "收回": "收回",
    "收回借款": "收回",
    "收回借出": "收回",
    "repayin": "收回",
    "还出": "偿还",
    "归还": "偿还",
    "还款": "偿还",
    "归还借款": "偿还",
    "repayout": "偿还",
}

ALLOWED_TYPES = {"收入", "支出", "借入", "贷出", "收回", "偿还"}
LOAN_TYPES = {"借入", "贷出", "收回", "偿还"}

DEFAULT_ALLOWED_CATEGORIES = [
    "吃喝",
    "日用消耗",
    "交通",
    "娱乐电子",
    "医疗",
    "人情往来",
    "给妈妈",
    "和女同志",
    "工资",
    "公积金",
    "房租",
]

HELPER_CODE_OPENPYXL = r"""
from __future__ import annotations

import json
import sys
from datetime import date, datetime, time as dt_time, timedelta, timezone
from pathlib import Path
from openpyxl import load_workbook

EXPECTED_HEADERS = ["ID", "Date", "Time", "Timezone", "Amount", "Currency", "Type", "Category", "Note", "PaymentChannel", "Status"]
LEGACY_HEADERS = ["ID", "Date", "Time", "Timezone", "Amount", "Currency", "Type", "Category", "Note", "Status"]
EARLY_LEGACY_HEADERS = ["ID", "Date", "Time", "Amount", "Currency", "Type", "Category", "Note", "Status"]
ID_COL = EXPECTED_HEADERS.index("ID") + 1
DATE_COL = EXPECTED_HEADERS.index("Date") + 1
TIME_COL = EXPECTED_HEADERS.index("Time") + 1
TIMEZONE_COL = EXPECTED_HEADERS.index("Timezone") + 1
PAYMENT_CHANNEL_COL = EXPECTED_HEADERS.index("PaymentChannel") + 1
STATUS_COL = EXPECTED_HEADERS.index("Status") + 1


def row_values(worksheet, row: int) -> list[object]:
    return [worksheet.cell(row=row, column=col).value for col in range(1, len(EXPECTED_HEADERS) + 1)]


def json_safe(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dt_time):
        return value.isoformat()
    return value


def find_last_non_empty_row(worksheet) -> int | None:
    for row in range(max(worksheet.max_row, 2), 1, -1):
        values = row_values(worksheet, row)
        if any(value not in (None, "") for value in values):
            return row
    return None


def find_row_by_id(worksheet, record_id: str) -> int:
    if not record_id:
        raise ValueError("ID 不能为空")
    
    last_row = worksheet.max_row
    for row in range(2, last_row + 1):
        if str(worksheet.cell(row=row, column=ID_COL).value) == str(record_id):
            return row
    raise ValueError(f"未找到 ID 为 {record_id} 的记录")


def read_headers(worksheet, count: int) -> list[object]:
    return [worksheet.cell(row=1, column=i).value for i in range(1, count + 1)]


def ensure_headers(worksheet):
    actual_headers = read_headers(worksheet, len(EXPECTED_HEADERS))
    if actual_headers == EXPECTED_HEADERS:
        return

    legacy_headers = read_headers(worksheet, len(LEGACY_HEADERS))
    early_legacy_headers = read_headers(worksheet, len(EARLY_LEGACY_HEADERS))
    if legacy_headers == LEGACY_HEADERS:
        worksheet.insert_cols(PAYMENT_CHANNEL_COL, amount=1)
        worksheet.cell(row=1, column=PAYMENT_CHANNEL_COL, value="PaymentChannel")
        return

    if early_legacy_headers == EARLY_LEGACY_HEADERS:
        worksheet.insert_cols(TIMEZONE_COL, amount=1)
        worksheet.cell(row=1, column=TIMEZONE_COL, value="Timezone")
        worksheet.insert_cols(PAYMENT_CHANNEL_COL, amount=1)
        worksheet.cell(row=1, column=PAYMENT_CHANNEL_COL, value="PaymentChannel")
        return

    if actual_headers[:2] == [None, None] or actual_headers[0] != "ID":
        for col, header in enumerate(EXPECTED_HEADERS, start=1):
            worksheet.cell(row=1, column=col, value=header)
        return

    raise ValueError(f"Header mismatch: {actual_headers}")


def normalize_timezone(value) -> str:
    if value in (None, ""):
        return "UTC+08:00"
    text = str(value).strip().upper().replace("GMT", "UTC")
    if text in {"UTC", "Z"}:
        return "UTC+00:00"
    if not text.startswith("UTC"):
        return "UTC+08:00"
    text = text.replace(" ", "")
    sign = text[3:4]
    if sign not in {"+", "-"}:
        return "UTC+08:00"
    rest = text[4:]
    if ":" in rest:
        hours_text, minutes_text = rest.split(":", 1)
    else:
        hours_text, minutes_text = rest, "00"
    try:
        hours = int(hours_text)
        minutes = int(minutes_text)
    except:
        return "UTC+08:00"
    return f"UTC{sign}{hours:02d}:{minutes:02d}"


def timezone_to_tzinfo(value) -> timezone:
    text = normalize_timezone(value)
    sign = 1 if text[3] == "+" else -1
    hours = int(text[4:6])
    minutes = int(text[7:9])
    return timezone(sign * timedelta(hours=hours, minutes=minutes))


def sort_worksheet(worksheet):
    # 简单的冒泡或提取重写排序。对于 Excel 脚本，提取所有数据排序后再写回最稳妥。
    rows = []
    last_row = find_last_non_empty_row(worksheet)
    if not last_row or last_row < 2:
        return

    for r in range(2, last_row + 1):
        rows.append(row_values(worksheet, r))

    def sort_key(row):
        d = row[DATE_COL-1]
        t = row[TIME_COL-1]
        tz_value = row[TIMEZONE_COL-1]

        if isinstance(d, str):
            try:
                d_obj = datetime.strptime(d, "%Y-%m-%d").date()
            except:
                d_obj = date(1970, 1, 1)
        elif isinstance(d, datetime):
            d_obj = d.date()
        else:
            d_obj = d or date(1970, 1, 1)

        if isinstance(t, str):
            try:
                t_obj = datetime.strptime(t, "%H:%M").time()
            except:
                t_obj = dt_time(0, 0)
        elif isinstance(t, datetime):
            t_obj = t.time()
        else:
            t_obj = t or dt_time(0, 0)

        local_dt = datetime.combine(d_obj, t_obj, tzinfo=timezone_to_tzinfo(tz_value))
        return local_dt.astimezone(timezone.utc)

    rows.sort(key=sort_key)

    for i, row_data in enumerate(rows, start=2):
        for j, val in enumerate(row_data, start=1):
            cell = worksheet.cell(row=i, column=j)
            cell.value = val
            if j == DATE_COL:
                cell.number_format = "yyyy-mm-dd"
            elif j == TIME_COL:
                cell.number_format = "hh:mm"


def main() -> int:
    excel_path = Path(sys.argv[1])
    payload_path = Path(sys.argv[2])

    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    action = payload.get("action", "append")
    record = payload.get("record")
    sheet_name = payload.get("sheet_name")

    workbook = load_workbook(excel_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]

    ensure_headers(worksheet)

    if action == "append":
        last_row = find_last_non_empty_row(worksheet) or 1
        next_row = last_row + 1
        for col, header in enumerate(EXPECTED_HEADERS, start=1):
            worksheet.cell(row=next_row, column=col, value=record[header])

        worksheet.cell(row=next_row, column=DATE_COL).number_format = "yyyy-mm-dd"
        worksheet.cell(row=next_row, column=TIME_COL).number_format = "hh:mm"
        
        # 排序
        sort_worksheet(worksheet)
        workbook.save(excel_path)

        # 排序后行号会变，需要重新找回该 ID 的行号返回给用户
        final_row = find_row_by_id(worksheet, record["ID"])
        print(json.dumps({"row": final_row}, ensure_ascii=False))
        return 0

    if action == "invalidate_id":
        target_id = payload.get("id")
        target_row = find_row_by_id(worksheet, target_id)
        worksheet.cell(row=target_row, column=STATUS_COL, value="作废")
        workbook.save(excel_path)
        print(json.dumps({"row": target_row}, ensure_ascii=False))
        return 0

    if action == "read_by_id":
        target_id = payload.get("id")
        target_row = find_row_by_id(worksheet, target_id)
        record_data = {key: json_safe(value) for key, value in zip(EXPECTED_HEADERS, row_values(worksheet, target_row))}
        print(json.dumps({"row": target_row, "record": record_data}, ensure_ascii=False))
        return 0

    if action == "invalidate_last":
        # 依然支持 invalidate_last，逻辑为找最后一行
        target_row = find_last_non_empty_row(worksheet)
        if not target_row or target_row < 2:
            raise ValueError("没有可作废的记录")
        worksheet.cell(row=target_row, column=STATUS_COL, value="作废")
        workbook.save(excel_path)
        print(json.dumps({"row": target_row}, ensure_ascii=False))
        return 0

    raise ValueError(f"Unsupported action: {action}")


if __name__ == "__main__":
    raise SystemExit(main())
"""

# win32com 暂不支持，因为 Linux/WSL 环境通常使用 openpyxl
# 这里保持简单，如果用户非要 win32com，会报错或需要类似实现
HELPER_CODE_WIN32COM = HELPER_CODE_OPENPYXL


def _pick_value(record: Dict[str, Any], field: str) -> Any:
    for alias in FIELD_ALIASES[field]:
        if alias in record and record[alias] is not None:
            return record[alias]
    return None


def _clean_config_list(values: Any) -> list[str]:
    if not isinstance(values, list):
        return []
    return [str(item).strip() for item in values if str(item).strip()]


def load_bot_config(config_path: Path = DEFAULT_CONFIG_PATH) -> Dict[str, Any]:
    if not config_path.exists():
        return {}
    return json.loads(config_path.read_text(encoding="utf-8"))


def get_allowed_categories(config: Dict[str, Any] | None = None, config_path: Path = DEFAULT_CONFIG_PATH) -> set[str]:
    loaded_config = config if config is not None else load_bot_config(config_path)
    categories = _clean_config_list(loaded_config.get("allowed_categories"))
    return set(categories or DEFAULT_ALLOWED_CATEGORIES)


def get_payment_channels(config: Dict[str, Any] | None = None, config_path: Path = DEFAULT_CONFIG_PATH) -> list[str]:
    loaded_config = config if config is not None else load_bot_config(config_path)
    return _clean_config_list(loaded_config.get("payment_channels"))


def get_default_payment_channel(config: Dict[str, Any] | None = None, config_path: Path = DEFAULT_CONFIG_PATH) -> str:
    loaded_config = config if config is not None else load_bot_config(config_path)
    default_channel = str(loaded_config.get("default_payment_channel", "")).strip()
    if not default_channel:
        return ""
    if default_channel not in get_payment_channels(config=loaded_config):
        return ""
    return default_channel


def normalize_status(raw_value: Any) -> str:
    if raw_value in (None, "", False, 0, "0", "否"):
        return STATUS_NORMAL
    if raw_value in (True, 1, "1", "是"):
        return STATUS_PENDING
    text = str(raw_value).strip()
    return text if text in ALLOWED_STATUS else STATUS_NORMAL


def normalize_timezone(raw_value: Any) -> str:
    if raw_value in (None, ""):
        return DEFAULT_TIMEZONE
    text = str(raw_value).strip()
    if text.upper() in {"UTC", "GMT", "Z"}:
        return "UTC+00:00"
    match = TIMEZONE_PATTERN.fullmatch(text)
    if not match:
        raise ValueError(f"时区非法: {raw_value}")

    sign, hours_text, minutes_text = match.groups()
    hours = int(hours_text)
    minutes = int(minutes_text or "00")
    if minutes >= 60:
        raise ValueError(f"时区非法: {raw_value}")

    total_minutes = hours * 60 + minutes
    if sign == "-":
        total_minutes = -total_minutes

    if total_minutes < -12 * 60 or total_minutes > 14 * 60:
        raise ValueError(f"时区超出范围: {raw_value}")
    if total_minutes in {-12 * 60, 14 * 60} and minutes != 0:
        raise ValueError(f"时区超出范围: {raw_value}")

    normalized_sign = "+" if total_minutes >= 0 else "-"
    abs_minutes = abs(total_minutes)
    normalized_hours, normalized_minutes = divmod(abs_minutes, 60)
    return f"UTC{normalized_sign}{normalized_hours:02d}:{normalized_minutes:02d}"


def timezone_to_tzinfo(raw_value: Any) -> timezone:
    normalized = normalize_timezone(raw_value)
    sign = 1 if normalized[3] == "+" else -1
    hours = int(normalized[4:6])
    minutes = int(normalized[7:9])
    return timezone(sign * timedelta(hours=hours, minutes=minutes))


def convert_telegram_timestamp(raw_value: Any, timezone_text: Any = DEFAULT_TIMEZONE) -> datetime:
    target_tz = timezone_to_tzinfo(timezone_text)
    if raw_value in (None, ""):
        return datetime.now(target_tz)
    if isinstance(raw_value, (int, float)):
        return datetime.fromtimestamp(float(raw_value), tz=timezone.utc).astimezone(target_tz)

    text = str(raw_value).strip()
    if not text:
        return datetime.now(target_tz)
    if re.fullmatch(r"\d+(\.\d+)?", text):
        return datetime.fromtimestamp(float(text), tz=timezone.utc).astimezone(target_tz)

    parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=target_tz)
    return parsed.astimezone(target_tz)


def normalize_record(record: Dict[str, Any]) -> Dict[str, Any]:
    normalized: Dict[str, Any] = {}
    for field in EXPECTED_HEADERS:
        value = _pick_value(record, field)
        if field not in ("Status", "Note", "PaymentChannel") and value is None:
             # ID 必须有
             if field == "ID":
                 raise ValueError("缺少必填字段: ID")
             raise ValueError(f"缺少必填字段: {field}")
        normalized[field] = value if value is not None else ""

    try:
        normalized["Amount"] = float(normalized["Amount"])
    except:
        raise ValueError(f"金额非法: {normalized['Amount']}")

    normalized["Timezone"] = normalize_timezone(normalized.get("Timezone"))
    normalized["Status"] = normalize_status(normalized.get("Status"))
    return normalized


def _windows_path(path: Path) -> str:
    completed = subprocess.run(["wslpath", "-w", str(path)], check=True, capture_output=True, text=True)
    return completed.stdout.strip()


def _run_excel_helper_with_result(excel_path: Path, payload: Dict[str, Any], backend: str) -> Dict[str, Any]:
    helper_code = HELPER_CODE_OPENPYXL # 强制使用 openpyxl 以支持新逻辑
    HELPER_TEMP_DIR.mkdir(parents=True, exist_ok=True)
    helper_path = HELPER_TEMP_DIR / "excel_append_helper.py"
    payload_path = HELPER_TEMP_DIR / "record.json"

    helper_path.write_text(helper_code, encoding="utf-8")
    payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    helper_win = _windows_path(helper_path)
    payload_win = _windows_path(payload_path)
    excel_win = _windows_path(excel_path)

    command = f"python '{helper_win}' '{excel_win}' '{payload_win}'"
    completed = subprocess.run(["powershell.exe", "-Command", command], capture_output=True, text=True)

    if completed.returncode != 0:
        raise RuntimeError(f"Excel 操作失败: {completed.stderr.strip() or completed.stdout.strip()}")

    return json.loads(completed.stdout.strip().splitlines()[-1])


def append_record_to_excel(excel_path: str | Path, record: Dict[str, Any], sheet_name: str | None = None, backend: str = "openpyxl") -> int:
    excel_path = Path(excel_path).resolve()
    payload = {"action": "append", "record": normalize_record(record), "sheet_name": sheet_name}
    result = _run_excel_helper_with_result(excel_path, payload, backend)
    return int(result["row"])


def invalidate_record_by_id(excel_path: str | Path, record_id: str, sheet_name: str | None = None, backend: str = "openpyxl") -> int:
    excel_path = Path(excel_path).resolve()
    payload = {"action": "invalidate_id", "id": record_id, "sheet_name": sheet_name}
    result = _run_excel_helper_with_result(excel_path, payload, backend)
    return int(result["row"])


def read_record_by_id(excel_path: str | Path, record_id: str, sheet_name: str | None = None, backend: str = "openpyxl") -> Dict[str, Any]:
    excel_path = Path(excel_path).resolve()
    payload = {"action": "read_by_id", "id": record_id, "sheet_name": sheet_name}
    result = _run_excel_helper_with_result(excel_path, payload, backend)
    return result["record"]


def invalidate_last_record_in_excel(excel_path: str | Path, sheet_name: str | None = None, backend: str = "openpyxl") -> int:
    excel_path = Path(excel_path).resolve()
    payload = {"action": "invalidate_last", "sheet_name": sheet_name}
    result = _run_excel_helper_with_result(excel_path, payload, backend)
    return int(result["row"])


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel-path", default=str(DEFAULT_EXCEL_PATH))
    parser.add_argument("--sheet-name")
    parser.add_argument("--json")
    parser.add_argument("--backend", default="openpyxl")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    if not args.json:
        print("Usage: --json '{\"ID\":\"...\", \"Date\":\"...\", ...}'")
        return 1
    
    record = json.loads(args.json)
    row_num = append_record_to_excel(args.excel_path, record, args.sheet_name, args.backend)
    print(json.dumps({"ok": True, "row": row_num}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
