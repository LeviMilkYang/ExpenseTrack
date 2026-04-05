
from __future__ import annotations

import json
import sys
from datetime import date, datetime, time as dt_time, timedelta, timezone
from pathlib import Path
from openpyxl import load_workbook

EXPECTED_HEADERS = ["ID", "Date", "Time", "Timezone", "Amount", "Currency", "Type", "Category", "Note", "PaymentChannel", "Status"]
LEGACY_HEADERS = ["ID", "Date", "Time", "Timezone", "Amount", "Currency", "Type", "Category", "Note", "Status"]
EARLY_LEGACY_HEADERS = ["ID", "Date", "Time", "Amount", "Currency", "Type", "Category", "Note", "Status"]
ID_COL = EXPECTED_HEADERS.index("ID") + 1
DATE_COL = EXPECTED_HEADERS.index("Date") + 1
TIME_COL = EXPECTED_HEADERS.index("Time") + 1
TIMEZONE_COL = EXPECTED_HEADERS.index("Timezone") + 1
PAYMENT_CHANNEL_COL = EXPECTED_HEADERS.index("PaymentChannel") + 1
STATUS_COL = EXPECTED_HEADERS.index("Status") + 1


def row_values(worksheet, row: int) -> list[object]:
    return [worksheet.cell(row=row, column=col).value for col in range(1, len(EXPECTED_HEADERS) + 1)]


def json_safe(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dt_time):
        return value.isoformat()
    return value


def find_last_non_empty_row(worksheet) -> int | None:
    for row in range(max(worksheet.max_row, 2), 1, -1):
        values = row_values(worksheet, row)
        if any(value not in (None, "") for value in values):
            return row
    return None


def find_row_by_id(worksheet, record_id: str) -> int:
    if not record_id:
        raise ValueError("ID 不能为空")
    
    last_row = worksheet.max_row
    for row in range(2, last_row + 1):
        if str(worksheet.cell(row=row, column=ID_COL).value) == str(record_id):
            return row
    raise ValueError(f"未找到 ID 为 {record_id} 的记录")


def read_headers(worksheet, count: int) -> list[object]:
    return [worksheet.cell(row=1, column=i).value for i in range(1, count + 1)]


def ensure_headers(worksheet):
    actual_headers = read_headers(worksheet, len(EXPECTED_HEADERS))
    if actual_headers == EXPECTED_HEADERS:
        return

    legacy_headers = read_headers(worksheet, len(LEGACY_HEADERS))
    early_legacy_headers = read_headers(worksheet, len(EARLY_LEGACY_HEADERS))
    if legacy_headers == LEGACY_HEADERS:
        worksheet.insert_cols(PAYMENT_CHANNEL_COL, amount=1)
        worksheet.cell(row=1, column=PAYMENT_CHANNEL_COL, value="PaymentChannel")
        return

    if early_legacy_headers == EARLY_LEGACY_HEADERS:
        worksheet.insert_cols(TIMEZONE_COL, amount=1)
        worksheet.cell(row=1, column=TIMEZONE_COL, value="Timezone")
        worksheet.insert_cols(PAYMENT_CHANNEL_COL, amount=1)
        worksheet.cell(row=1, column=PAYMENT_CHANNEL_COL, value="PaymentChannel")
        return

    if actual_headers[:2] == [None, None] or actual_headers[0] != "ID":
        for col, header in enumerate(EXPECTED_HEADERS, start=1):
            worksheet.cell(row=1, column=col, value=header)
        return

    raise ValueError(f"Header mismatch: {actual_headers}")


def normalize_timezone(value) -> str:
    if value in (None, ""):
        return "UTC+08:00"
    text = str(value).strip().upper().replace("GMT", "UTC")
    if text in {"UTC", "Z"}:
        return "UTC+00:00"
    if not text.startswith("UTC"):
        return "UTC+08:00"
    text = text.replace(" ", "")
    sign = text[3:4]
    if sign not in {"+", "-"}:
        return "UTC+08:00"
    rest = text[4:]
    if ":" in rest:
        hours_text, minutes_text = rest.split(":", 1)
    else:
        hours_text, minutes_text = rest, "00"
    try:
        hours = int(hours_text)
        minutes = int(minutes_text)
    except:
        return "UTC+08:00"
    return f"UTC{sign}{hours:02d}:{minutes:02d}"


def timezone_to_tzinfo(value) -> timezone:
    text = normalize_timezone(value)
    sign = 1 if text[3] == "+" else -1
    hours = int(text[4:6])
    minutes = int(text[7:9])
    return timezone(sign * timedelta(hours=hours, minutes=minutes))


def row_utc_datetime(row) -> datetime:
    d = row[DATE_COL - 1]
    t = row[TIME_COL - 1]
    tz_value = row[TIMEZONE_COL - 1]

    if isinstance(d, str):
        try:
            d_obj = datetime.strptime(d, "%Y-%m-%d").date()
        except:
            d_obj = date(1970, 1, 1)
    elif isinstance(d, datetime):
        d_obj = d.date()
    else:
        d_obj = d or date(1970, 1, 1)

    if isinstance(t, str):
        try:
            t_obj = datetime.strptime(t, "%H:%M").time()
        except:
            t_obj = dt_time(0, 0)
    elif isinstance(t, datetime):
        t_obj = t.time()
    else:
        t_obj = t or dt_time(0, 0)

    local_dt = datetime.combine(d_obj, t_obj, tzinfo=timezone_to_tzinfo(tz_value))
    return local_dt.astimezone(timezone.utc)


def find_insert_row(worksheet, record) -> int:
    target_dt = row_utc_datetime([record[header] for header in EXPECTED_HEADERS])
    last_row = find_last_non_empty_row(worksheet)
    if not last_row or last_row < 2:
        return 2

    for row in range(last_row, 1, -1):
        values = row_values(worksheet, row)
        if not any(value not in (None, "") for value in values):
            continue
        if row_utc_datetime(values) <= target_dt:
            return row + 1
    return 2


def write_record(worksheet, target_row: int, record) -> None:
    for col, header in enumerate(EXPECTED_HEADERS, start=1):
        worksheet.cell(row=target_row, column=col, value=record[header])

    worksheet.cell(row=target_row, column=DATE_COL).number_format = "yyyy-mm-dd"
    worksheet.cell(row=target_row, column=TIME_COL).number_format = "hh:mm"


def main() -> int:
    excel_path = Path(sys.argv[1])
    payload_path = Path(sys.argv[2])

    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    action = payload.get("action", "append")
    record = payload.get("record")
    sheet_name = payload.get("sheet_name")

    workbook = load_workbook(excel_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]

    ensure_headers(worksheet)

    if action == "append":
        insert_row = find_insert_row(worksheet, record)
        worksheet.insert_rows(insert_row, amount=1)
        write_record(worksheet, insert_row, record)
        workbook.save(excel_path)
        print(json.dumps({"row": insert_row}, ensure_ascii=False))
        return 0

    if action == "invalidate_id":
        target_id = payload.get("id")
        target_row = find_row_by_id(worksheet, target_id)
        worksheet.cell(row=target_row, column=STATUS_COL, value="作废")
        workbook.save(excel_path)
        print(json.dumps({"row": target_row}, ensure_ascii=False))
        return 0

    if action == "read_by_id":
        target_id = payload.get("id")
        target_row = find_row_by_id(worksheet, target_id)
        record_data = {key: json_safe(value) for key, value in zip(EXPECTED_HEADERS, row_values(worksheet, target_row))}
        print(json.dumps({"row": target_row, "record": record_data}, ensure_ascii=False))
        return 0

    if action == "invalidate_last":
        # 依然支持 invalidate_last，逻辑为找最后一行
        target_row = find_last_non_empty_row(worksheet)
        if not target_row or target_row < 2:
            raise ValueError("没有可作废的记录")
        worksheet.cell(row=target_row, column=STATUS_COL, value="作废")
        workbook.save(excel_path)
        print(json.dumps({"row": target_row}, ensure_ascii=False))
        return 0

    raise ValueError(f"Unsupported action: {action}")


if __name__ == "__main__":
    raise SystemExit(main())
