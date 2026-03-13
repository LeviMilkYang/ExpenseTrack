from __future__ import annotations

import argparse
import socket
from collections import defaultdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import json
import os
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any, Dict

from openpyxl import load_workbook

from append_excel_entry import (
    EXPECTED_HEADERS,
    append_record_to_excel,
    invalidate_last_record_in_excel,
    invalidate_record_in_excel,
    normalize_record,
    read_record_from_excel,
)
from generate_expense_report import _load_records, refresh_report_workbook

API_TIMEOUT_SECONDS = 60
BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
RESTART_SCRIPT_PATH = BASE_DIR / "restart_bot.sh"
BUDGET_SHEET_NAME = "预算"
VOID_MARK = "作废"


def get_monthly_file_path(base_name: str, ext: str) -> Path:
    """返回如 logs/2026-03.log 或 indexes/2026-03.json 的路径"""
    month_str = time.strftime("%Y-%m", time.localtime())
    folder = BASE_DIR / base_name
    folder.mkdir(parents=True, exist_ok=True)
    return folder / f"{month_str}.{ext}"


def log_json(payload: Dict[str, Any]) -> None:
    # 注入当前时间戳 (YYYY-MM-DD HH:MM:SS)
    payload["timestamp"] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    # 交互式运行时输出到 stdout；后台运行时只写文件，避免与 nohup 重定向重复落盘。
    if sys.stdout.isatty():
        print(json.dumps(payload, ensure_ascii=False), flush=True)

    log_file = get_monthly_file_path("logs", "log")
    with log_file.open("a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False) + "\n")


def api_request(token: str, method: str, payload: Dict[str, Any] | None = None, timeout: int = 30) -> Dict[str, Any]:
    url = f"https://api.telegram.org/bot{token}/{method}"
    data = None
    headers = {}
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"

    request = urllib.request.Request(url, data=data, headers=headers, method="POST" if payload is not None else "GET")
    with urllib.request.urlopen(request, timeout=timeout) as response:
        body = response.read().decode("utf-8")
    result = json.loads(body)
    if not result.get("ok"):
        raise RuntimeError(f"Telegram API {method} failed: {result}")
    return result


def _to_budget_decimal(value: Any) -> Decimal:
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError(f"预算金额非法: {value!r}") from exc
    if amount < 0:
        raise ValueError(f"预算金额不能为负数: {value!r}")
    return amount


def load_budget_config(excel_path: Path) -> Dict[str, Dict[str, Any]]:
    workbook = load_workbook(excel_path, data_only=False)
    if BUDGET_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"{excel_path.name} 中缺少“{BUDGET_SHEET_NAME}”sheet")

    worksheet = workbook[BUDGET_SHEET_NAME]
    headers = [worksheet.cell(row=1, column=idx).value for idx in range(1, 4)]
    if headers != ["Category", "Amount", "Fixed"]:
        raise ValueError(f"预算sheet表头不正确: {headers}")

    categories: Dict[str, Dict[str, Any]] = {}
    for row in worksheet.iter_rows(min_row=2, max_col=3, values_only=True):
        category_value, amount_value, fixed_value = row
        if all(value in (None, "") for value in row):
            continue

        category = str(category_value).strip()
        if not category:
            raise ValueError("预算sheet存在空分类")
        categories[category] = {
            "amount": _to_budget_decimal(amount_value),
            "fixed": str(fixed_value).strip().lower() in {"1", "true", "yes", "y", "fixed", "是"} if fixed_value not in (None, "") else False,
        }

    if not categories:
        raise ValueError("预算sheet中没有有效预算配置")
    return categories


def resolve_budget_period(envelope: Dict[str, Any]) -> str:
    timestamp = envelope.get("telegram_timestamp")
    if timestamp:
        return time.strftime("%Y-%m", time.localtime(float(timestamp)))
    return time.strftime("%Y-%m", time.localtime())


def format_amount(value: Decimal) -> str:
    quantized = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if quantized == quantized.to_integral():
        return str(int(quantized))
    return format(quantized.normalize(), "f")


def build_budget_reply(envelope: Dict[str, Any], excel_path: Path) -> str:
    budget_config = load_budget_config(excel_path)
    period = resolve_budget_period(envelope)

    spent_by_category: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    unbudgeted_spent: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))

    for record in _load_records(excel_path):
        if record.currency != "CNY" or record.record_type != "支出":
            continue
        if record.record_date.strftime("%Y-%m") != period:
            continue

        if record.category in budget_config:
            spent_by_category[record.category] += record.amount
        else:
            unbudgeted_spent[record.category] += record.amount

    lines = [f"{period} 剩余预算"]
    total_budget = Decimal("0")
    total_spent = Decimal("0")

    for category, config in budget_config.items():
        if config["fixed"]:
            continue

        budget_amount = config["amount"]
        spent_amount = spent_by_category.get(category, Decimal("0"))
        remaining_amount = budget_amount - spent_amount
        total_budget += budget_amount
        total_spent += spent_amount

        lines.append(
            f"{category}：剩余 {format_amount(remaining_amount)} / 预算 {format_amount(budget_amount)}（已用 {format_amount(spent_amount)}）"
        )

    total_remaining = total_budget - total_spent
    lines.append(f"合计：剩余 {format_amount(total_remaining)} / 预算 {format_amount(total_budget)}（已用 {format_amount(total_spent)}）")

    if unbudgeted_spent:
        extras = "，".join(
            f"{category} {format_amount(amount)}"
            for category, amount in sorted(unbudgeted_spent.items(), key=lambda item: (-item[1], item[0]))
        )
        lines.append(f"未设预算支出：{extras}")

    lines.append("固定支出房租、给妈妈不计入本指令显示。")
    return "\n".join(lines)


def classify_network_error(exc: BaseException) -> str:
    if isinstance(exc, urllib.error.HTTPError):
        return f"http_{exc.code}"
    if isinstance(exc, TimeoutError):
        return "timeout"
    if isinstance(exc, socket.timeout):
        return "timeout"

    error_text = str(exc).lower()
    if "timed out" in error_text:
        return "timeout"
    if "eof occurred in violation of protocol" in error_text:
        return "ssl_eof"
    if "remote end closed connection without response" in error_text:
        return "remote_close"
    if "handshake operation timed out" in error_text:
        return "tls_handshake_timeout"
    return "unknown_network_error"


def load_state(state_path: Path) -> Dict[str, Any]:
    if not state_path.exists():
        return {"offset": 0, "token": ""}
    return json.loads(state_path.read_text(encoding="utf-8"))


def save_state(state_path: Path, state: Dict[str, Any]) -> None:
    state_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def load_message_index() -> Dict[str, Any]:
    index_path = get_monthly_file_path("indexes", "json")
    if not index_path.exists():
        return {}
    return json.loads(index_path.read_text(encoding="utf-8"))


def save_message_index(index: Dict[str, Any]) -> None:
    index_path = get_monthly_file_path("indexes", "json")
    index_path.write_text(json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")


def configured_project_dir(state: Dict[str, Any]) -> Path:
    raw_value = str(state.get("project_dir", "")).strip()
    if raw_value:
        return Path(raw_value).expanduser()
    return PROJECT_DIR


def configured_excel_path(state: Dict[str, Any]) -> Path:
    return configured_project_dir(state) / "expense.xlsx"


def configured_allowed_username(state: Dict[str, Any]) -> str:
    return str(state.get("allowed_username", "")).strip()


def current_report_period() -> str:
    return time.strftime("%Y-%m", time.localtime())


def refresh_report_if_period_changed(state_path: Path, excel_path: Path, verbose: bool) -> None:
    state = load_state(state_path)
    current_period = current_report_period()
    last_period = str(state.get("report_period", "")).strip()
    report_path = excel_path.with_name("expense_report.xlsx")

    if last_period == current_period and report_path.exists():
        return

    refresh_report_workbook(excel_path, report_path)
    state["report_period"] = current_period
    save_state(state_path, state)
    if verbose:
        log_json({"stage": "report_refreshed", "period": current_period, "report_path": str(report_path)})


def load_token(state_path: Path, legacy_token_path: Path | None = None) -> str:
    state = load_state(state_path)
    token = str(state.get("token", "")).strip()
    if token:
        return token

    if legacy_token_path is not None and legacy_token_path.exists():
        legacy_token = legacy_token_path.read_text(encoding="utf-8").strip()
        if legacy_token:
            state["token"] = legacy_token
            save_state(state_path, state)
            return legacy_token

    raise ValueError("telegram_bot_state.json 中缺少 token")


def queue_restart_confirmation(state_path: Path, chat_id: int, reply_to_message_id: int) -> None:
    state = load_state(state_path)
    state["pending_restart_notice"] = {
        "chat_id": chat_id,
        "reply_to_message_id": reply_to_message_id,
        "created_at": int(time.time()),
    }
    save_state(state_path, state)


def send_pending_restart_confirmation(token: str, state_path: Path, verbose: bool) -> None:
    state = load_state(state_path)
    notice = state.get("pending_restart_notice")
    if not isinstance(notice, dict):
        return

    chat_id = notice.get("chat_id")
    reply_to_message_id = notice.get("reply_to_message_id")
    if not isinstance(chat_id, int) or not isinstance(reply_to_message_id, int):
        state.pop("pending_restart_notice", None)
        save_state(state_path, state)
        return

    send_reply(token, chat_id, reply_to_message_id, "机器人已重启完成")
    state.pop("pending_restart_notice", None)
    save_state(state_path, state)
    if verbose:
        log_json({"stage": "restart_confirmed", "chat_id": chat_id, "reply_to_message_id": reply_to_message_id})


def build_message_reference(message: Dict[str, Any] | None) -> Dict[str, Any] | None:
    if not isinstance(message, dict):
        return None

    chat = message.get("chat", {}) or {}
    return {
        "message_id": message.get("message_id"),
        "chat_id": chat.get("id"),
        "sender": (message.get("from", {}) or {}).get("username") or (message.get("from", {}) or {}).get("first_name") or "",
        "telegram_timestamp": message.get("date"),
        "text": message.get("text") or "",
    }


def build_envelope(message: Dict[str, Any]) -> Dict[str, Any]:
    sender = message.get("from", {})
    chat = message.get("chat", {})
    return {
        "message_id": message.get("message_id"),
        "chat_id": chat.get("id"),
        "sender": sender.get("username") or sender.get("first_name") or "",
        "telegram_timestamp": message.get("date"),
        "text": message.get("text") or "",
        "reply_to_message": build_message_reference(message.get("reply_to_message")),
    }


def should_process(message: Dict[str, Any], allowed_username: str) -> tuple[bool, str]:
    sender = message.get("from", {}) or {}
    username = sender.get("username")
    if not allowed_username:
        return False, "whitelist username is not configured"
    if username != allowed_username:
        return False, f"忽略非白名单用户: {username!r}"

    text = (message.get("text") or "").strip()
    if not text:
        return False, "忽略非文本消息"
    return True, ""


def message_index_key(chat_id: int, message_id: int) -> str:
    return f"{chat_id}:{message_id}"


def normalize_index_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        number = float(value)
        return int(number) if number.is_integer() else number
    return str(value).strip()


def build_record_fingerprint(record: Dict[str, Any]) -> str:
    stable_record = {field: normalize_index_value(record.get(field)) for field in EXPECTED_HEADERS}
    return json.dumps(stable_record, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def register_record_mapping(envelope: Dict[str, Any], result: Dict[str, Any]) -> None:
    chat_id = envelope.get("chat_id")
    message_id = envelope.get("message_id")
    row = result.get("row")
    sheet_name = result.get("sheet_name")
    record = result.get("record")
    if not isinstance(chat_id, int) or not isinstance(message_id, int):
        return
    if not isinstance(row, int) or not isinstance(sheet_name, str) or not sheet_name.strip():
        return
    if not isinstance(record, dict):
        return

    index = load_message_index()
    index[message_index_key(chat_id, message_id)] = {
        "chat_id": chat_id,
        "message_id": message_id,
        "sheet_name": sheet_name,
        "row": row,
        "record_fingerprint": build_record_fingerprint(record),
        "voided": False,
        "created_at": int(time.time()),
    }
    save_message_index(index)


def invalidate_reply_target(
    envelope: Dict[str, Any],
    excel_path: Path,
    backend: str,
) -> Dict[str, Any]:
    reply_to_message = envelope.get("reply_to_message")
    if not isinstance(reply_to_message, dict):
        raise ValueError("当前消息没有引用历史消息")

    chat_id = reply_to_message.get("chat_id")
    message_id = reply_to_message.get("message_id")
    if not isinstance(chat_id, int) or not isinstance(message_id, int):
        raise ValueError("引用消息缺少有效的 message_id")

    index = load_message_index()
    key = message_index_key(chat_id, message_id)
    entry = index.get(key)
    if not isinstance(entry, dict):
        raise ValueError("这条消息没有对应账本记录")
    if entry.get("voided") is True:
        raise ValueError("该记录已作废")

    sheet_name = str(entry.get("sheet_name", "")).strip()
    row = entry.get("row")
    if not sheet_name or not isinstance(row, int):
        raise ValueError("索引记录不完整，无法执行作废")

    current_record = read_record_from_excel(excel_path, row=row, sheet_name=sheet_name, backend=backend)
    if str(current_record.get("NeedConfirm", "")).strip() == VOID_MARK:
        entry["voided"] = True
        entry["voided_at"] = int(time.time())
        save_message_index(index)
        raise ValueError("该记录已作废")

    current_fingerprint = build_record_fingerprint(current_record)
    expected_fingerprint = str(entry.get("record_fingerprint", ""))
    if current_fingerprint != expected_fingerprint:
        raise ValueError("记录已变化，需要人工确认")

    invalidated_row = invalidate_record_in_excel(excel_path, row=row, sheet_name=sheet_name, backend=backend)
    entry["voided"] = True
    entry["voided_at"] = int(time.time())
    save_message_index(index)
    return {"row": invalidated_row, "sheet_name": sheet_name, "message_id": message_id}


def run_bridge_prompt(workdir: Path, envelope: Dict[str, Any]) -> str:
    completed = subprocess.run(
        [
            "python3",
            str(workdir / "telegram_codex_bridge.py"),
            "prompt",
            "--json",
            json.dumps(envelope, ensure_ascii=False),
        ],
        cwd=workdir,
        capture_output=True,
        text=True,
        check=True,
    )
    output = completed.stdout.strip()
    if not output:
        raise RuntimeError("bridge prompt 未返回内容")
    return output


def run_gemini(workdir: Path, prompt: str) -> Dict[str, Any]:
    env = os.environ.copy()
    # 清理 IDE 相关变量以避免无头模式下的 IDEClient 错误
    for key in ["GEMINI_CLI_IDE_SERVER_PORT", "GEMINI_CLI_IDE_AUTH_TOKEN"]:
        env.pop(key, None)

    max_retries = 3
    last_error = ""
    for attempt in range(max_retries):
        completed = subprocess.run(
            [
                "gemini",
                "--prompt",
                prompt,
                "--output-format",
                "json",
            ],
            cwd=workdir,
            capture_output=True,
            text=True,
            env=env,
        )
        if completed.returncode == 0:
            try:
                outer_json = json.loads(completed.stdout.strip())
                raw_response = outer_json.get("response", "").strip()
                
                if raw_response.startswith("```json"):
                    raw_response = raw_response[7:].strip()
                elif raw_response.startswith("```"):
                    raw_response = raw_response[3:].strip()
                if raw_response.endswith("```"):
                    raw_response = raw_response[:-3].strip()
                    
                return json.loads(raw_response)
            except (json.JSONDecodeError, KeyError) as e:
                import re
                match = re.search(r"(\{.*\})", raw_response, re.DOTALL)
                if match:
                    try:
                        return json.loads(match.group(1))
                    except json.JSONDecodeError:
                        pass
                last_error = f"Failed to parse gemini response: {e}"
        else:
            last_error = (completed.stderr or completed.stdout).strip() or "gemini call failed"
        
        if attempt < max_retries - 1:
            time.sleep(2 ** attempt)
            
    raise RuntimeError(last_error)


def run_bridge_apply(workdir: Path, envelope: Dict[str, Any], gemini_output: Dict[str, Any], backend: str, excel_path: Path) -> Dict[str, Any]:
    payload = dict(envelope)
    payload["gemini_output"] = gemini_output
    completed = subprocess.run(
        [
            "python3",
            str(workdir / "telegram_codex_bridge.py"),
            "apply",
            "--backend",
            backend,
            "--excel-path",
            str(excel_path),
            "--json",
            json.dumps(payload, ensure_ascii=False),
        ],
        cwd=workdir,
        capture_output=True,
        text=True,
        check=True,
    )
    return json.loads(completed.stdout.strip())


def send_reply(token: str, chat_id: int, reply_to_message_id: int, text: str) -> Dict[str, Any]:
    return api_request(
        token,
        "sendMessage",
        {
            "chat_id": chat_id,
            "text": text,
            "reply_to_message_id": reply_to_message_id,
        },
    )


def trigger_bot_restart(verbose: bool) -> None:
    subprocess.Popen(
        ["bash", str(RESTART_SCRIPT_PATH)],
        cwd=BASE_DIR,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        stdin=subprocess.DEVNULL,
        start_new_session=True,
    )
    if verbose:
        log_json({"stage": "restart_triggered", "script": str(RESTART_SCRIPT_PATH)})


def get_fallback_record(envelope: Dict[str, Any]) -> Dict[str, Any]:
    """当 AI 失效时，生成一条兜底记录"""
    ts = envelope.get("telegram_timestamp")
    if ts:
        local_tm = time.localtime(float(ts))
        date_str = time.strftime("%Y-%m-%d", local_tm)
        time_str = time.strftime("%H:%M", local_tm)
    else:
        date_str = time.strftime("%Y-%m-%d")
        time_str = time.strftime("%H:%M")

    return {
        "Date": date_str,
        "Time": time_str,
        "Amount": 0,
        "Currency": "CNY",
        "Type": "支出",
        "Category": "未分类",
        "Note": f"[AI失败兜底] {envelope['text']}",
        "NeedConfirm": True,
    }


def handle_message(
    token: str,
    workdir: Path,
    state_path: Path,
    message: Dict[str, Any],
    backend: str,
    excel_path: Path,
    verbose: bool,
) -> None:
    envelope = build_envelope(message)
    if envelope["text"] == "重启":
        queue_restart_confirmation(state_path, envelope["chat_id"], envelope["message_id"])
        send_reply(token, envelope["chat_id"], envelope["message_id"], "正在重启机器人")
        trigger_bot_restart(verbose)
        return

    if envelope["text"] == "预算":
        send_reply(token, envelope["chat_id"], envelope["message_id"], "正在处理预算...")
        try:
            reply = build_budget_reply(envelope, excel_path)
            if verbose:
                log_json(
                    {
                        "stage": "budget_viewed",
                        "message_id": envelope["message_id"],
                        "period": resolve_budget_period(envelope),
                    }
                )
        except Exception as exc:
            if verbose:
                log_json({"stage": "budget_failed", "message_id": envelope["message_id"], "error": str(exc)})
            reply = f"预算查询失败：{exc}"
        send_reply(token, envelope["chat_id"], envelope["message_id"], reply)
        return

    if envelope["text"] == "作废":
        if envelope.get("reply_to_message"):
            result = invalidate_reply_target(envelope, excel_path, backend)
            if verbose:
                log_json({"stage": "invalidated_reply", "message_id": envelope["message_id"], "result": result})
            send_reply(
                token,
                envelope["chat_id"],
                envelope["message_id"],
                f"已作废对应消息的记录：第 {result['row']} 行（{result['sheet_name']}）",
            )
            return

        row = invalidate_last_record_in_excel(excel_path, backend=backend)
        if verbose:
            log_json({"stage": "invalidated_last", "message_id": envelope["message_id"], "row": row})
        send_reply(token, envelope["chat_id"], envelope["message_id"], f"已作废：第 {row} 行")
        return

    # 发送“正在处理”消息
    send_reply(token, envelope["chat_id"], envelope["message_id"], "正在处理...")

    try:
        try:
            prompt = run_bridge_prompt(workdir, envelope)
            gemini_output = run_gemini(workdir, prompt)
            result = run_bridge_apply(workdir, envelope, gemini_output, backend, excel_path)
            fallback_used = False
        except Exception as ai_exc:
            if verbose:
                log_json({"stage": "ai_failed_using_fallback", "message_id": envelope["message_id"], "error": str(ai_exc)})
            
            record = get_fallback_record(envelope)
            normalized = normalize_record(record)
            sheet_name = record["Date"].split("-")[0]
            row = append_record_to_excel(excel_path, normalized, sheet_name, backend)
            
            result = {
                "ok": True,
                "record": normalized,
                "sheet_name": sheet_name,
                "row": row,
                "fallback": True,
                "error_detail": str(ai_exc)
            }
            fallback_used = True

        if verbose:
            log_json({"stage": "applied", "message_id": envelope["message_id"], "result": result})

        if result.get("ignored"):
            reply = f"已忽略：{result.get('reason', '不是记账相关消息')}"
            send_reply(token, envelope["chat_id"], envelope["message_id"], reply)
            return

        register_record_mapping(envelope, result)

        note = result["record"]["Note"]
        amount = result["record"]["Amount"]
        record_type = result["record"]["Type"]
        category = result["record"]["Category"]
        
        if fallback_used:
            reply = f"⚠️ AI 处理失败，已为您自动记录原文：\n{record_type} / {category} / {amount}\n备注：{note}\n请稍后手动核对（第 {result['row']} 行）"
        else:
            confirm_hint = "，待确认" if result["record"]["NeedConfirm"] else ""
            reply = f"已记账：{record_type} / {category} / {amount}"
            if note:
                reply += f" / {note}"
            reply += f"，第 {result['row']} 行{confirm_hint}"
        
        send_reply(token, envelope["chat_id"], envelope["message_id"], reply)
    except Exception as exc:
        error_text = f"系统故障，无法记账：{exc}"
        if verbose:
            log_json({"stage": "critical_error", "message_id": message.get("message_id"), "error": str(exc)})
        send_reply(token, message["chat"]["id"], message["message_id"], error_text)


def poll_loop(
    token: str,
    workdir: Path,
    state_path: Path,
    allowed_username: str,
    backend: str,
    excel_path: Path,
    verbose: bool,
) -> None:
    state = load_state(state_path)
    offset = int(state.get("offset", 0))

    while True:
        try:
            refresh_report_if_period_changed(state_path, excel_path, verbose)
            state = load_state(state_path)
            state["offset"] = offset
            result = api_request(
                token,
                "getUpdates",
                {
                    "offset": offset,
                    "timeout": API_TIMEOUT_SECONDS,
                    "allowed_updates": ["message"],
                },
                timeout=API_TIMEOUT_SECONDS + 10,
            )
            for update in result.get("result", []):
                update_id = int(update["update_id"])
                offset = update_id + 1
                state = load_state(state_path)
                state["offset"] = offset
                save_state(state_path, state)

                message = update.get("message")
                if not message:
                    continue

                accepted, reason = should_process(message, allowed_username)
                if verbose:
                    preview = (message.get("text") or "")[:40]
                    log_json({"stage": "received", "accepted": accepted, "reason": reason, "text": preview})

                if not accepted:
                    continue

                try:
                    handle_message(token, workdir, state_path, message, backend, excel_path, verbose)
                except Exception as exc:
                    log_json({"stage": "handle_message_crash", "error": str(exc)})
        except urllib.error.HTTPError as exc:
            if exc.code == 401:
                if verbose:
                    log_json(
                        {
                            "stage": "fatal_auth_error",
                            "error_type": classify_network_error(exc),
                            "error": f"HTTP Error {exc.code}: {exc.reason}",
                        }
                    )
                raise RuntimeError("Telegram bot token unauthorized; exiting daemon") from exc
            if verbose:
                log_json({"stage": "network_error", "error_type": classify_network_error(exc), "error": str(exc)})
            time.sleep(5)
        except urllib.error.URLError as exc:
            if verbose:
                log_json({"stage": "network_error", "error_type": classify_network_error(exc), "error": str(exc)})
            time.sleep(5)
        except Exception as exc:
            if verbose:
                log_json({"stage": "loop_error", "error_type": classify_network_error(exc), "error": str(exc)})
            time.sleep(5)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Telegram expense bot daemon.")
    parser.add_argument("--state-file", default=str(BASE_DIR / "telegram_bot_state.json"))
    parser.add_argument("--legacy-token-file", default=str(BASE_DIR / "bot_token.txt"))
    parser.add_argument("--excel-path", default="")
    parser.add_argument("--backend", choices=["win32com", "openpyxl"], default="openpyxl")
    parser.add_argument("--once", action="store_true")
    parser.add_argument("--verbose", action="store_true")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    workdir = BASE_DIR
    state_path = Path(args.state_file)
    state = load_state(state_path)
    excel_path = Path(args.excel_path) if str(args.excel_path).strip() else configured_excel_path(state)
    allowed_username = configured_allowed_username(state)
    token = load_token(state_path, Path(args.legacy_token_file))

    me = api_request(token, "getMe")
    if args.verbose:
        log_json({"stage": "getMe", "result": me["result"]})

    send_pending_restart_confirmation(token, state_path, args.verbose)

    if args.once:
        log_json({"ok": True, "bot": me["result"], "allowed_username_configured": bool(allowed_username), "excel_path": str(excel_path)})
        return 0

    refresh_report_if_period_changed(state_path, excel_path, args.verbose)
    poll_loop(token, workdir, state_path, allowed_username, args.backend, excel_path, args.verbose)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
